import openpyxl
from openpyxl import Workbook
import random

wb = Workbook()
wb = openpyxl.load_workbook('game_database.xlsx')

'''Создание новых листов'''
ttl = ['Mero', 'Svechki', 'NiP']
for title in ttl:
    if title not in wb.sheetnames:
        wb.create_sheet(title)

ws = wb.active
# wb.save('game_database.xlsx')


def get_random_game():
    rc = random.randint(1, 9)
    return (f"{ws.cell(row=rc, column=4).value}: \n"
            f"\n"
            f" {ws.cell(row=rc, column=5).value}")


def get_igra_minutka():
    ws = wb['Games']
    while True:
        rc = random.randint(1, ws.max_row + 1)
        if ws.cell(row=rc, column=2).value == 'Игра-минутка':
            return (f"{ws.cell(row=rc, column=4).value}: \n"
                    f"\n"
                    f" {ws.cell(row=rc, column=5).value}")


def get_igra_na_lidera():
    ws = wb['Games']
    while True:
        rc = random.randint(1, ws.max_row + 1)
        if ws.cell(row=rc, column=2).value == 'На выявление лидера':
            return (f"{ws.cell(row=rc, column=4).value}: \n"
                    f"\n"
                    f" {ws.cell(row=rc, column=5).value}")


def get_igra_na_znakomstvo():
    ws = wb['Games']
    while True:
        rc = random.randint(1, ws.max_row + 1)
        if ws.cell(row=rc, column=2).value == 'На знакомство':
            return (f"{ws.cell(row=rc, column=4).value}: \n"
                    f"\n"
                    f" {ws.cell(row=rc, column=5).value}")


def nip_pooschreniya():
    ws = wb['NiP']
    while True:
        rc = random.randint(1, ws.max_row + 1)
        if ws.cell(row=rc, column=1).value == 'Поощрения':
            return ws.cell(row=rc, column=2).value


def nip_nakazaniya():
    ws = wb['NiP']
    while True:
        rc = random.randint(1, ws.max_row + 1)
        if ws.cell(row=rc, column=1).value == 'Наказания':
            return ws.cell(row=rc, column=2).value
